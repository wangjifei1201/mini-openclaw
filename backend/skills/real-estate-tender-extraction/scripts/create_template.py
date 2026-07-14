from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT = BASE_DIR / "templates" / "地产招标文件核心信息提取模板.xlsx"

SHEETS = {
    "招标关键信息汇总": [
        "序号", "业务类别", "关联类别", "信息主题", "关键信息", "原文摘录", "来源文件", "页码", "章节/条款号",
        "业务解读", "对投标文件的影响", "风险等级", "可信度", "是否需人工复核", "复核意见", "备注",
    ],
    "综合通用事项": [
        "序号", "事项类型", "信息主题", "关键信息", "原文摘录", "来源位置", "时间/金额/数量", "责任部门",
        "对应投标动作", "风险等级", "是否需人工复核", "备注",
    ],
    "商务标": [
        "序号", "商务事项类型", "信息主题", "报价/合同/计价要求", "原文摘录", "来源位置", "涉及金额/费率/公式",
        "对报价文件的影响", "风险等级", "是否需人工复核", "备注",
    ],
    "技术标": [
        "序号", "技术事项类型", "信息主题", "技术编写要求", "原文摘录", "来源位置", "评分分值/响应要求",
        "对技术方案的影响", "暗标相关", "风险等级", "是否需人工复核", "备注",
    ],
    "资信标": [
        "序号", "资信事项类型", "信息主题", "资格/证明/评分要求", "原文摘录", "来源位置", "证明材料要求",
        "对资信文件的影响", "风险等级", "是否需人工复核", "备注",
    ],
    "风险与复核清单": [
        "序号", "风险类型", "涉及业务类别", "风险描述", "原文摘录", "来源位置", "影响范围", "建议复核动作",
        "优先级", "责任人", "复核结论", "备注",
    ],
    "字段字典": ["字段", "填写口径", "枚举/示例"],
}

DICTIONARY_ROWS = [
    ("业务类别", "主业务分类", "综合通用、商务标、技术标、资信标"),
    ("关联类别", "与该事项相关的其他类别", "商务标；技术标；资信标"),
    ("信息主题", "简短概括条款主题", "投标截止时间、报价方式、项目经理资格"),
    ("关键信息", "不改变原意的结构化提炼", "不得替代原文摘录"),
    ("原文摘录", "逐字复制的原文证据", "不得改写、润色、合并创造"),
    ("来源文件", "信息所在文件", "招标文件、答疑、补遗、合同条款、技术标准"),
    ("页码", "文件页码或 PDF 页码", "缺少页码时可信度不得为高"),
    ("章节/条款号", "章节名称、条款号、表格编号或附件名称", "投标人须知前附表 3.4.1"),
    ("业务解读", "面向投标编制的解释", "与原文摘录分列"),
    ("风险等级", "风险程度", "高、中、低"),
    ("可信度", "来源定位和信息明确程度", "高、中、低"),
    ("是否需人工复核", "是否需要业务人员复核", "是、否"),
    ("风险类型", "风险分类", "废标、资格、报价、合同、时间、格式、暗标、冲突、缺失、其他"),
]

VALIDATIONS = {
    "业务类别": "综合通用,商务标,技术标,资信标",
    "风险等级": "高,中,低",
    "可信度": "高,中,低",
    "是否需人工复核": "是,否",
    "优先级": "高,中,低",
    "暗标相关": "是,否",
    "风险类型": "废标,资格,报价,合同,时间,格式,暗标,冲突,缺失,其他",
}

WIDTHS = {
    "序号": 8,
    "业务类别": 14,
    "关联类别": 16,
    "信息主题": 24,
    "关键信息": 36,
    "原文摘录": 60,
    "来源文件": 22,
    "页码": 10,
    "章节/条款号": 24,
    "业务解读": 42,
    "对投标文件的影响": 42,
    "复核意见": 24,
    "备注": 24,
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")


def apply_validation(ws, header: str, options: str) -> None:
    headers = [cell.value for cell in ws[1]]
    if header not in headers:
        return
    col_idx = headers.index(header) + 1
    col = get_column_letter(col_idx)
    dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}500")


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=500):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        header = column_cells[0].value
        width = WIDTHS.get(header, 18)
        ws.column_dimensions[column_cells[0].column_letter].width = width
    ws.row_dimensions[1].height = 28


def create_workbook() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for name, headers in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        if name == "字段字典":
            for row in DICTIONARY_ROWS:
                ws.append(row)
        else:
            ws.append(["示例/待填写"] + ["" for _ in headers[1:]])
            for cell in ws[2]:
                cell.fill = NOTE_FILL
        style_sheet(ws)
        for header, options in VALIDATIONS.items():
            apply_validation(ws, header, options)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    create_workbook()
