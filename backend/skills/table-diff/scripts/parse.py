#!/usr/bin/env python3
"""
file-parser: 解析表格文件，提取结构元信息和数据内容
支持格式：.xlsx, .csv
"""

import csv
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


def detect_encoding(file_path: str) -> str:
    """检测 CSV 文件编码，UTF-8 优先，失败回退 GBK"""
    for enc in ["utf-8", "gbk", "gb2312", "latin-1"]:
        try:
            with open(file_path, "r", encoding=enc) as f:
                f.read(1024)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "utf-8"


def infer_dtype(values: list) -> str:
    """推断列数据类型"""
    non_null = [v for v in values if v is not None and v != ""]
    if not non_null:
        return "string"

    # 尝试 integer：所有值必须是纯整数（不含小数点，或小数部分为零）
    try:
        all_int = True
        for v in non_null:
            s = str(v)
            # 含小数点的，检查小数部分是否为零
            if "." in s:
                if float(s) != int(float(s)):
                    all_int = False
                    break
            # 不含小数点的，必须是合法整数
            elif not s.lstrip("-").isdigit():
                all_int = False
                break
        if all_int:
            return "integer"
    except (ValueError, TypeError):
        pass

    # 尝试 float
    try:
        for v in non_null:
            float(str(v))
        return "float"
    except (ValueError, TypeError):
        pass

    # 尝试 datetime（简单判断）
    import re
    dt_pattern = re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}")
    if all(dt_pattern.match(str(v)) for v in non_null):
        return "datetime"

    return "string"


def parse_xlsx(file_path: str, sheet_name: str = None, max_rows: int = 1000) -> dict:
    """解析 .xlsx 文件"""
    if openpyxl is None:
        return {"error": "dependency_missing", "message": "openpyxl 未安装，请执行 pip install openpyxl"}

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    if sheet_name:
        if sheet_name not in sheets:
            wb.close()
            return {"error": "sheet_not_found", "message": f"Sheet '{sheet_name}' 不存在", "available_sheets": sheets}
        ws = wb[sheet_name]
    else:
        ws = wb[sheets[0]]
        sheet_name = sheets[0]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {"error": "empty_file", "message": "文件内容为空"}

    # 第一行作为表头
    headers = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(rows[0])]
    data_rows = rows[1:max_rows]

    # 跳过全空行
    data_rows = [r for r in data_rows if any(v is not None for v in r)]

    # 按列组织数据
    col_data = {h: [] for h in headers}
    for row in data_rows:
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            col_data[h].append(val)

    # 构建列元信息
    columns_meta = []
    for i, h in enumerate(headers):
        values = col_data[h]
        non_null = [v for v in values if v is not None and v != ""]
        null_count = len(values) - len(non_null)
        unique_vals = list(set(str(v) for v in non_null))
        sample_values = unique_vals[:3]

        columns_meta.append({
            "name": h,
            "index": i,
            "dtype": infer_dtype(non_null),
            "null_count": null_count,
            "unique_count": len(unique_vals),
            "sample_values": sample_values
        })

    # 构建数据行
    data = []
    for row in data_rows:
        row_dict = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            row_dict[h] = val
        data.append(row_dict)

    return {
        "meta": {
            "source_file": os.path.basename(file_path),
            "sheet_name": sheet_name,
            "row_count": len(data_rows),
            "col_count": len(headers),
            "columns": columns_meta
        },
        "data": data
    }


def parse_csv(file_path: str, max_rows: int = 1000) -> dict:
    """解析 .csv 文件"""
    encoding = detect_encoding(file_path)

    with open(file_path, "r", encoding=encoding) as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return {"error": "empty_file", "message": "文件内容为空"}

    headers = [str(h).strip() if h.strip() else f"column_{i}" for i, h in enumerate(rows[0])]
    data_rows = rows[1:max_rows + 1]

    # 跳过全空行
    data_rows = [r for r in data_rows if any(v.strip() for v in r if v)]

    # 按列组织数据
    col_data = {h: [] for h in headers}
    for row in data_rows:
        for i, h in enumerate(headers):
            val = row[i].strip() if i < len(row) else None
            # 数值转换
            if val == "" or val is None:
                val = None
            else:
                try:
                    if "." in val:
                        val = float(val)
                    else:
                        val = int(val)
                except ValueError:
                    pass
            col_data[h].append(val)

    # 构建列元信息
    columns_meta = []
    for i, h in enumerate(headers):
        values = col_data[h]
        non_null = [v for v in values if v is not None and v != ""]
        null_count = len(values) - len(non_null)
        unique_vals = list(set(str(v) for v in non_null))
        sample_values = unique_vals[:3]

        columns_meta.append({
            "name": h,
            "index": i,
            "dtype": infer_dtype(non_null),
            "null_count": null_count,
            "unique_count": len(unique_vals),
            "sample_values": sample_values
        })

    # 构建数据行
    data = []
    for row in data_rows:
        row_dict = {}
        for i, h in enumerate(headers):
            val = row[i].strip() if i < len(row) else None
            if val == "":
                val = None
            else:
                try:
                    if "." in val:
                        val = float(val)
                    else:
                        val = int(val)
                except ValueError:
                    pass
            row_dict[h] = val
        data.append(row_dict)

    return {
        "meta": {
            "source_file": os.path.basename(file_path),
            "sheet_name": None,
            "row_count": len(data_rows),
            "col_count": len(headers),
            "columns": columns_meta
        },
        "data": data
    }


def parse(file_path: str, sheet_name: str = None, max_rows: int = 1000) -> dict:
    """统一入口"""
    if not os.path.exists(file_path):
        return {"error": "file_not_found", "message": f"文件不存在: {file_path}"}

    ext = Path(file_path).suffix.lower()
    if ext == ".xlsx":
        return parse_xlsx(file_path, sheet_name, max_rows)
    elif ext == ".csv":
        return parse_csv(file_path, max_rows)
    else:
        return {"error": "unsupported_format", "message": f"不支持的格式: {ext}，仅支持 .xlsx 和 .csv"}


if __name__ == "__main__":
    # CLI 调用方式：python parse.py <file_path> [--sheet <name>] [--max-rows <n>]
    import argparse

    parser = argparse.ArgumentParser(description="表格文件解析器")
    parser.add_argument("file_path", help="文件路径")
    parser.add_argument("--sheet", default=None, help="Sheet 名称（仅 xlsx）")
    parser.add_argument("--max-rows", type=int, default=1000, help="最大读取行数")
    parser.add_argument("--output", default=None, help="输出文件路径，不指定则输出到 stdout")

    args = parser.parse_args()

    result = parse(args.file_path, args.sheet, args.max_rows)

    output = json.dumps(result, ensure_ascii=False, indent=2, default=str)

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(output)
    else:
        print(output)
