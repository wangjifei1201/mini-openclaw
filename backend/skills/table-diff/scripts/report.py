#!/usr/bin/env python3
"""
diff-reporter: 差异报告生成器
支持 HTML 和 Excel 两种格式
"""

import csv
import html
import json


def escape_html(value) -> str:
    """HTML 安全展示。"""
    if value is None:
        return ""
    return html.escape(str(value), quote=True)


def markdown_escape(value) -> str:
    """Markdown 表格单元格安全展示。"""
    if value is None:
        return ""
    return (
        html.escape(str(value), quote=True)
        .replace("|", "\\|")
        .replace("\r", "<br>")
        .replace("\n", "<br>")
    )


def spreadsheet_safe_value(value) -> str:
    """Prevent CSV/Excel formula execution for untrusted cell values."""
    if value is None:
        return ""
    text = str(value)
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def stable_json(value) -> str:
    """Stable readable JSON serialization for row data."""
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def csv_value(value) -> str:
    """CSV 展示并防止公式注入。"""
    return spreadsheet_safe_value(value)


def primary_key_label(primary_key: dict) -> str:
    """主键字典展示为 key=value / key=value。"""
    return " / ".join(f"{k}={v}" for k, v in primary_key.items())

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# ============================================================
# HTML 报告
# ============================================================

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>表格比对报告</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif; background: #f5f7fa; color: #333; padding: 24px; }}
.container {{ max-width: 1200px; margin: 0 auto; }}
h1 {{ font-size: 22px; font-weight: 600; margin-bottom: 20px; color: #1a1a1a; }}
h2 {{ font-size: 16px; font-weight: 600; margin: 20px 0 12px; color: #1a1a1a; }}

/* 概览卡片 */
.overview {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; margin-bottom: 24px; }}
.card {{ background: #fff; border-radius: 8px; padding: 16px; text-align: center; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
.card .label {{ font-size: 12px; color: #888; margin-bottom: 4px; }}
.card .value {{ font-size: 24px; font-weight: 700; }}
.card .value.accent {{ color: #e63946; }}
.card .value.green {{ color: #2a9d8f; }}
.card .value.blue {{ color: #457b9d; }}

/* 列级摘要 */
.col-summary {{ background: #fff; border-radius: 8px; padding: 16px; margin-bottom: 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
.col-summary table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
.col-summary th {{ text-align: left; padding: 8px 12px; border-bottom: 2px solid #eee; color: #666; font-weight: 500; }}
.col-summary td {{ padding: 6px 12px; border-bottom: 1px solid #f0f0f0; }}
.bar {{ display: inline-block; height: 8px; border-radius: 4px; background: #e63946; vertical-align: middle; }}

/* 筛选 */
.filters {{ margin-bottom: 16px; display: flex; gap: 8px; }}
.filters button {{ padding: 6px 16px; border: 1px solid #ddd; border-radius: 6px; background: #fff; cursor: pointer; font-size: 13px; }}
.filters button.active {{ background: #457b9d; color: #fff; border-color: #457b9d; }}

/* 差异明细 */
.diff-table {{ background: #fff; border-radius: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
.diff-table table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
.diff-table th {{ background: #f8f9fa; padding: 10px 12px; text-align: left; border-bottom: 2px solid #eee; font-weight: 500; color: #555; position: sticky; top: 0; }}
.diff-table td {{ padding: 8px 12px; border-bottom: 1px solid #f0f0f0; }}
.diff-table tr.hidden {{ display: none; }}

.row-left-only {{ background: #fff5f5; }}
.row-right-only {{ background: #f0fff4; }}
.cell-changed-old {{ background: #ffe0e0; }}
.cell-changed-new {{ background: #d4edda; }}

.type-badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 500; }}
.badge-changed {{ background: #fff3cd; color: #856404; }}
.badge-left {{ background: #ffe0e0; color: #c0392b; }}
.badge-right {{ background: #d4edda; color: #155724; }}

.collapsed .diff-row {{ display: none; }}
.collapsed .diff-row:first-child {{ display: table-row; }}
.expand-hint {{ text-align: center; padding: 12px; color: #888; font-size: 13px; cursor: pointer; }}
</style>
</head>
<body>
<div class="container">
<h1>表格比对报告</h1>

{overview_html}
{col_summary_html}

<h2>差异明细</h2>
<div class="filters">
  <button class="active" onclick="filter(event, 'all')">全部</button>
  <button onclick="filter(event, 'value_changed')">值变化</button>
  <button onclick="filter(event, 'left_only')">仅左表</button>
  <button onclick="filter(event, 'right_only')">仅右表</button>
</div>
<div class="diff-table">
  {detail_html}
</div>

</div>
<script>
function filter(event, type) {{
  document.querySelectorAll('.filters button').forEach(b => b.classList.remove('active'));
  event.target.classList.add('active');
  document.querySelectorAll('.diff-row').forEach(row => {{
    if (type === 'all' || row.dataset.type === type) {{
      row.style.display = '';
    }} else {{
      row.style.display = 'none';
    }}
  }});
}}
</script>
</body>
</html>"""


def generate_overview_html(summary: dict) -> str:
    """生成概览区 HTML"""
    rate_pct = f"{summary['change_rate'] * 100:.1f}%"
    return f"""
<div class="overview">
  <div class="card"><div class="label">左表行数</div><div class="value blue">{escape_html(summary['total_left'])}</div></div>
  <div class="card"><div class="label">右表行数</div><div class="value blue">{escape_html(summary['total_right'])}</div></div>
  <div class="card"><div class="label">匹配行数</div><div class="value">{escape_html(summary['matched'])}</div></div>
  <div class="card"><div class="label">值变化行</div><div class="value accent">{escape_html(summary['value_changed'])}</div></div>
  <div class="card"><div class="label">仅左表</div><div class="value accent">{escape_html(summary['left_only'])}</div></div>
  <div class="card"><div class="label">仅右表</div><div class="value green">{escape_html(summary['right_only'])}</div></div>
  <div class="card"><div class="label">未变更行</div><div class="value">{escape_html(summary['unchanged'])}</div></div>
  <div class="card"><div class="label">变更率</div><div class="value accent">{escape_html(rate_pct)}</div></div>
</div>"""


def generate_col_summary_html(col_diff: dict) -> str:
    """生成列级摘要 HTML"""
    if not col_diff:
        return ""

    rows = ""
    for col, info in sorted(col_diff.items(), key=lambda x: -x[1]["change_rate"]):
        rate_pct = f"{info['change_rate'] * 100:.1f}%"
        bar_width = max(4, int(info["change_rate"] * 200))
        rows += f"""
    <tr>
      <td>{escape_html(col)}</td>
      <td>{escape_html(info['changed_count'])}</td>
      <td>{escape_html(rate_pct)}</td>
      <td><span class="bar" style="width:{bar_width}px"></span></td>
    </tr>"""

    return f"""
<div class="col-summary">
  <h2>列级变更摘要</h2>
  <table>
    <tr><th>列名</th><th>变更次数</th><th>变更率</th><th>分布</th></tr>
    {rows}
  </table>
</div>"""


def _val_display(val) -> str:
    """值展示。"""
    if val is None:
        return "<em style='color:#999'>空</em>"
    return escape_html(val)


def generate_detail_html(diffs: list, primary_key_cols: list, all_compare_cols: list) -> str:
    """生成差异明细表 HTML"""
    # 表头：类型 | 主键列... | 左表值列... | 右表值列...
    headers = ["类型"] + primary_key_cols
    for col in all_compare_cols:
        headers.append(f"左表·{col}")
    for col in all_compare_cols:
        headers.append(f"右表·{col}")

    thead = "<tr>" + "".join(f"<th>{escape_html(h)}</th>" for h in headers) + "</tr>"

    tbody = ""
    for d in diffs:
        dtype = d["type"]
        pk = d["primary_key"]

        if dtype == "value_changed":
            badge = f'<span class="type-badge badge-changed">变化</span>'
            changes = {c["column"]: c for c in d["changes"]}
            changed_cols = set(changes.keys())

            left_cells = []
            right_cells = []
            for col in all_compare_cols:
                if col in changed_cols:
                    left_cells.append(f'<td class="cell-changed-old">{_val_display(changes[col]["left_value"])}</td>')
                    right_cells.append(f'<td class="cell-changed-new">{_val_display(changes[col]["right_value"])}</td>')
                else:
                    # 需要从原始数据取值，但 diffs 里没存 unchanged 的值
                    left_cells.append(f'<td>-</td>')
                    right_cells.append(f'<td>-</td>')

            pk_cells = "".join(f"<td>{_val_display(pk.get(c))}</td>" for c in primary_key_cols)
            tbody += f'<tr class="diff-row" data-type="value_changed"><td>{badge}</td>{pk_cells}{"".join(left_cells)}{"".join(right_cells)}</tr>'

        elif dtype == "left_only":
            badge = '<span class="type-badge badge-left">仅左表</span>'
            row_data = d.get("row_data", {})
            pk_cells = "".join(f"<td>{_val_display(pk.get(c))}</td>" for c in primary_key_cols)
            left_cells = "".join(f"<td>{_val_display(row_data.get(c))}</td>" for c in all_compare_cols)
            right_cells = "".join("<td>-</td>" for _ in all_compare_cols)
            tbody += f'<tr class="diff-row row-left-only" data-type="left_only"><td>{badge}</td>{pk_cells}{left_cells}{right_cells}</tr>'

        elif dtype == "right_only":
            badge = '<span class="type-badge badge-right">仅右表</span>'
            row_data = d.get("row_data", {})
            pk_cells = "".join(f"<td>{_val_display(pk.get(c))}</td>" for c in primary_key_cols)
            left_cells = "".join("<td>-</td>" for _ in all_compare_cols)
            right_cells = "".join(f"<td>{_val_display(row_data.get(c))}</td>" for c in all_compare_cols)
            tbody += f'<tr class="diff-row row-right-only" data-type="right_only"><td>{badge}</td>{pk_cells}{left_cells}{right_cells}</tr>'

    return f"<table><thead>{thead}</thead><tbody>{tbody}</tbody></table>"


def generate_html_report(diff_result: dict, left_meta: dict, right_meta: dict) -> str:
    """生成 HTML 报告，返回 HTML 字符串"""
    summary = diff_result["summary"]
    col_diff = diff_result.get("column_diff_summary", {})
    diffs = diff_result.get("diffs", [])

    # 推断主键列和比对列
    primary_key_cols = []
    if diffs:
        pk = diffs[0].get("primary_key", {})
        primary_key_cols = list(pk.keys())

    # 比对列 = 列级摘要里的列 + left_only/right_only row_data 里的列 - 主键列
    all_cols_set = set(col_diff.keys())
    for d in diffs:
        if d["type"] == "value_changed":
            all_cols_set.update(c["column"] for c in d.get("changes", []))
        elif d["type"] in ("left_only", "right_only"):
            all_cols_set.update(d.get("row_data", {}).keys())
    all_compare_cols = sorted(all_cols_set - set(primary_key_cols))

    overview_html = generate_overview_html(summary)
    col_summary_html = generate_col_summary_html(col_diff)
    detail_html = generate_detail_html(diffs, primary_key_cols, all_compare_cols)

    return HTML_TEMPLATE.format(
        overview_html=overview_html,
        col_summary_html=col_summary_html,
        detail_html=detail_html
    )



def generate_markdown_report(diff_result: dict, left_meta: dict, right_meta: dict) -> str:
    """生成 Markdown 报告。"""
    summary = diff_result["summary"]
    col_diff = diff_result.get("column_diff_summary", {})
    diffs = diff_result.get("diffs", [])
    change_rate = f"{summary['change_rate'] * 100:.1f}%"

    lines = [
        "# 表格比对报告",
        "",
        "## 基本信息",
        "",
        "| 项目 | 左表 | 右表 |",
        "|---|---:|---:|",
        f"| 文件 | {markdown_escape(left_meta.get('source_file', ''))} | {markdown_escape(right_meta.get('source_file', ''))} |",
        f"| 行数 | {summary['total_left']} | {summary['total_right']} |",
        "",
        "## 摘要",
        "",
        "| 指标 | 数值 |",
        "|---|---:|",
        f"| 匹配行数 | {summary['matched']} |",
        f"| 值变化行 | {summary['value_changed']} |",
        f"| 仅左表行 | {summary['left_only']} |",
        f"| 仅右表行 | {summary['right_only']} |",
        f"| 未变更行 | {summary['unchanged']} |",
        f"| 变更率 | {change_rate} |",
        "",
    ]

    if col_diff:
        lines.extend([
            "## 列级变更摘要",
            "",
            "| 列名 | 变更次数 | 变更率 |",
            "|---|---:|---:|",
        ])
        for col, info in sorted(col_diff.items(), key=lambda x: -x[1]["change_rate"]):
            lines.append(f"| {markdown_escape(col)} | {info['changed_count']} | {info['change_rate'] * 100:.1f}% |")
        lines.append("")

    lines.extend([
        "## 差异明细",
        "",
        "| 类型 | 主键 | 行号 | 列名 | 左表值 | 右表值 |",
        "|---|---|---:|---|---|---|",
    ])

    for item in diffs:
        dtype = item["type"]
        type_label = {"value_changed": "值变化", "left_only": "仅左表", "right_only": "仅右表"}.get(dtype, dtype)
        pk = markdown_escape(primary_key_label(item.get("primary_key", {})))
        row_number = markdown_escape(item.get("row_number", ""))

        if dtype == "value_changed":
            for change in item.get("changes", []):
                lines.append(
                    f"| {type_label} | {pk} | {row_number} | {markdown_escape(change['column'])} | {markdown_escape(change.get('left_value'))} | {markdown_escape(change.get('right_value'))} |"
                )
        elif dtype == "left_only":
            row_json = stable_json(item.get('row_data', {}))
            lines.append(f"| {type_label} | {pk} | {row_number} | 整行 | {markdown_escape(row_json)} |  |")
        elif dtype == "right_only":
            row_json = stable_json(item.get('row_data', {}))
            lines.append(f"| {type_label} | {pk} | {row_number} | 整行 |  | {markdown_escape(row_json)} |")

    lines.extend([
        "",
        "## 总结",
        "",
        f"本次比对匹配 {summary['matched']} 行，其中 {summary['value_changed']} 行存在值变化，{summary['left_only']} 行仅左表存在，{summary['right_only']} 行仅右表存在。",
    ])

    return "\n".join(lines)


# ============================================================
# Excel 报告
# ============================================================

def generate_excel_report(diff_result: dict, left_meta: dict, right_meta: dict, output_path: str) -> str:
    """生成 Excel 报告"""
    if openpyxl is None:
        return json.dumps({"error": "dependency_missing", "message": "openpyxl 未安装"})

    summary = diff_result["summary"]
    col_diff = diff_result.get("column_diff_summary", {})
    diffs = diff_result.get("diffs", [])

    wb = openpyxl.Workbook()

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="457B9D")
    red_fill = PatternFill("solid", fgColor="FFE0E0")
    green_fill = PatternFill("solid", fgColor="D4EDDA")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD")
    )

    def style_header(ws, col_count):
        for i in range(1, col_count + 1):
            cell = ws.cell(row=1, column=i)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # --- Sheet1: 概览 ---
    ws1 = wb.active
    ws1.title = "概览"
    overview_data = [
        ["指标", "数值"],
        ["左表文件", left_meta.get("source_file", "")],
        ["右表文件", right_meta.get("source_file", "")],
        ["左表行数", summary["total_left"]],
        ["右表行数", summary["total_right"]],
        ["匹配行数", summary["matched"]],
        ["值变化行", summary["value_changed"]],
        ["仅左表行", summary["left_only"]],
        ["仅右表行", summary["right_only"]],
        ["未变更行", summary["unchanged"]],
        ["变更率", f"{summary['change_rate'] * 100:.1f}%"],
    ]
    # 列级摘要追加
    if col_diff:
        overview_data.append([])
        overview_data.append(["列级变更摘要"])
        overview_data.append(["列名", "变更次数", "变更率"])
        for col, info in sorted(col_diff.items(), key=lambda x: -x[1]["change_rate"]):
            overview_data.append([col, info["changed_count"], f"{info['change_rate'] * 100:.1f}%"])

    for row in overview_data:
        ws1.append([spreadsheet_safe_value(value) for value in row])
    style_header(ws1, 3)
    auto_width(ws1)

    # --- Sheet2: 值变化明细 ---
    ws2 = wb.create_sheet("值变化明细")
    ws2.append(["主键", "列名", "旧值(左表)", "新值(右表)"])
    for d in diffs:
        if d["type"] == "value_changed":
            pk_str = " / ".join(str(v) for v in d["primary_key"].values())
            for change in d["changes"]:
                ws2.append([spreadsheet_safe_value(value) for value in [pk_str, change["column"], change["left_value"], change["right_value"]]])
                # 变化值标色
                row_idx = ws2.max_row
                ws2.cell(row=row_idx, column=3).fill = red_fill
                ws2.cell(row=row_idx, column=4).fill = green_fill
    style_header(ws2, 4)
    auto_width(ws2)

    # --- Sheet3: 新增行(右表独有) ---
    ws3 = wb.create_sheet("新增行")
    right_only_diffs = [d for d in diffs if d["type"] == "right_only"]
    if right_only_diffs:
        cols = list(right_only_diffs[0].get("primary_key", {}).keys()) + list(right_only_diffs[0].get("row_data", {}).keys())
        ws3.append([spreadsheet_safe_value(col) for col in cols])
        for d in right_only_diffs:
            row_vals = list(d["primary_key"].values()) + list(d["row_data"].values())
            ws3.append([spreadsheet_safe_value(v) for v in row_vals])
            for i in range(1, len(row_vals) + 1):
                ws3.cell(row=ws3.max_row, column=i).fill = green_fill
    style_header(ws3, max(len(cols), 1) if right_only_diffs else 1)
    auto_width(ws3)

    # --- Sheet4: 删除行(左表独有) ---
    ws4 = wb.create_sheet("删除行")
    left_only_diffs = [d for d in diffs if d["type"] == "left_only"]
    if left_only_diffs:
        cols = list(left_only_diffs[0].get("primary_key", {}).keys()) + list(left_only_diffs[0].get("row_data", {}).keys())
        ws4.append([spreadsheet_safe_value(col) for col in cols])
        for d in left_only_diffs:
            row_vals = list(d["primary_key"].values()) + list(d["row_data"].values())
            ws4.append([spreadsheet_safe_value(v) for v in row_vals])
            for i in range(1, len(row_vals) + 1):
                ws4.cell(row=ws4.max_row, column=i).fill = red_fill
    style_header(ws4, max(len(cols), 1) if left_only_diffs else 1)
    auto_width(ws4)

    # --- Sheet5: 完整对比 ---
    ws5 = wb.create_sheet("完整对比")
    # 推断主键列和比对列
    primary_key_cols = []
    if diffs:
        pk = diffs[0].get("primary_key", {})
        primary_key_cols = list(pk.keys())
    all_cols_set = set(col_diff.keys())
    for d in diffs:
        if d["type"] == "value_changed":
            all_cols_set.update(c["column"] for c in d.get("changes", []))
        elif d["type"] in ("left_only", "right_only"):
            all_cols_set.update(d.get("row_data", {}).keys())
    all_compare_cols = sorted(all_cols_set - set(primary_key_cols))

    headers = ["类型"] + primary_key_cols
    for col in all_compare_cols:
        headers.append(f"左·{col}")
    for col in all_compare_cols:
        headers.append(f"右·{col}")
    ws5.append([spreadsheet_safe_value(header) for header in headers])

    for d in diffs:
        dtype = d["type"]
        pk = d["primary_key"]
        type_label = {"value_changed": "变化", "left_only": "仅左表", "right_only": "仅右表"}.get(dtype, dtype)

        row = [type_label] + [pk.get(c, "") for c in primary_key_cols]

        if dtype == "value_changed":
            changes = {c["column"]: c for c in d["changes"]}
            for col in all_compare_cols:
                row.append(changes[col]["left_value"] if col in changes else "")
            for col in all_compare_cols:
                row.append(changes[col]["right_value"] if col in changes else "")
        elif dtype == "left_only":
            rd = d.get("row_data", {})
            for col in all_compare_cols:
                row.append(rd.get(col, ""))
            for _ in all_compare_cols:
                row.append("")
        elif dtype == "right_only":
            for _ in all_compare_cols:
                row.append("")
            rd = d.get("row_data", {})
            for col in all_compare_cols:
                row.append(rd.get(col, ""))

        row = [spreadsheet_safe_value(v) for v in row]
        ws5.append(row)

        # 标色
        r = ws5.max_row
        if dtype == "left_only":
            for i in range(1, len(row) + 1):
                ws5.cell(row=r, column=i).fill = red_fill
        elif dtype == "right_only":
            for i in range(1, len(row) + 1):
                ws5.cell(row=r, column=i).fill = green_fill
        elif dtype == "value_changed":
            changes = {c["column"]: c for c in d.get("changes", [])}
            for idx, col in enumerate(all_compare_cols):
                if col in changes:
                    left_col_idx = 1 + len(primary_key_cols) + idx
                    right_col_idx = 1 + len(primary_key_cols) + len(all_compare_cols) + idx
                    ws5.cell(row=r, column=left_col_idx).fill = red_fill
                    ws5.cell(row=r, column=right_col_idx).fill = green_fill

    style_header(ws5, len(headers))
    auto_width(ws5)

    wb.save(output_path)
    return output_path


def generate_csv_report(diff_result: dict, output_path: str) -> str:
    """生成机器可读 CSV 明细。"""
    diffs = diff_result.get("diffs", [])
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["type", "primary_key", "row_number", "column", "left_value", "right_value"])
        writer.writeheader()
        for item in diffs:
            dtype = item["type"]
            pk = csv_value(primary_key_label(item.get("primary_key", {})))
            row_number = csv_value(item.get("row_number", ""))
            if dtype == "value_changed":
                for change in item.get("changes", []):
                    writer.writerow({
                        "type": csv_value(dtype),
                        "primary_key": pk,
                        "row_number": row_number,
                        "column": csv_value(change.get("column")),
                        "left_value": csv_value(change.get("left_value")),
                        "right_value": csv_value(change.get("right_value")),
                    })
            elif dtype == "left_only":
                writer.writerow({
                    "type": csv_value(dtype),
                    "primary_key": pk,
                    "row_number": row_number,
                    "column": csv_value("__row__"),
                    "left_value": csv_value(stable_json(item.get("row_data", {}))),
                    "right_value": "",
                })
            elif dtype == "right_only":
                writer.writerow({
                    "type": csv_value(dtype),
                    "primary_key": pk,
                    "row_number": row_number,
                    "column": csv_value("__row__"),
                    "left_value": "",
                    "right_value": csv_value(stable_json(item.get("row_data", {}))),
                })
    return output_path


# ============================================================
# 统一入口
# ============================================================

def report(diff_result: dict, left_meta: dict, right_meta: dict,
           format: str = "html", output_path: str = None) -> str:
    """生成差异报告"""
    if not diff_result.get("summary"):
        return json.dumps({"error": "invalid_diff_result", "message": "差异结果格式非法"})

    if format == "html":
        html_text = generate_html_report(diff_result, left_meta, right_meta)
        if output_path:
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(html_text)
            return output_path
        return html_text

    elif format in ("excel", "xlsx"):
        if not output_path:
            return json.dumps({"error": "output_required", "message": "Excel 格式必须指定 output_path"})
        return generate_excel_report(diff_result, left_meta, right_meta, output_path)

    elif format in ("markdown", "md"):
        markdown_text = generate_markdown_report(diff_result, left_meta, right_meta)
        if output_path:
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(markdown_text)
            return output_path
        return markdown_text

    elif format == "csv":
        if not output_path:
            return json.dumps({"error": "output_required", "message": "CSV 格式必须指定 output_path"})
        return generate_csv_report(diff_result, output_path)

    else:
        return json.dumps({"error": "unsupported_format", "message": f"不支持的格式: {format}"})


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="差异报告生成器")
    parser.add_argument("diff_file", help="差异结果 JSON 文件路径")
    parser.add_argument("--left-meta", required=True, help="左表 meta JSON 文件路径")
    parser.add_argument("--right-meta", required=True, help="右表 meta JSON 文件路径")
    parser.add_argument("--format", choices=["html", "excel", "xlsx", "markdown", "md", "csv"], required=True, help="输出格式")
    parser.add_argument("--output", default=None, help="输出文件路径")

    args = parser.parse_args()
    if args.format in ("csv", "xlsx", "excel") and not args.output:
        parser.error("--output is required for csv/xlsx/excel formats")

    with open(args.diff_file, "r", encoding="utf-8") as f:
        diff_result = json.load(f)
    with open(args.left_meta, "r", encoding="utf-8") as f:
        left_meta = json.load(f)
    with open(args.right_meta, "r", encoding="utf-8") as f:
        right_meta = json.load(f)

    result = report(diff_result, left_meta, right_meta, args.format, args.output)
    if args.format in ("html", "markdown", "md") and not args.output:
        print(result)
    else:
        print(f"报告已生成: {result}")
